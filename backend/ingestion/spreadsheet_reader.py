"""Safe spreadsheet / CSV reading — no macros, no formula execution."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Any

from backend.ingestion.exceptions import (
    EmptyFileError,
    FormulaRejectedError,
    LimitExceededError,
    UnsupportedFormatError,
)
from backend.ingestion.limits import (
    MAX_CELL_CHARS,
    MAX_COLUMNS,
    MAX_CSV_LINES_SCAN,
    MAX_PREVIEW_ROWS,
    MAX_SHEETS,
)


_FORMULA_PREFIX = re.compile(r"^[=+\-@]")


@dataclass
class SheetPreview:
    name: str
    rows: list[list[Any]]
    empty: bool
    row_count_scanned: int
    col_count: int


@dataclass
class SpreadsheetContent:
    format: str
    sheets: list[SheetPreview]
    delimiter: str | None = None
    encoding: str | None = None
    warnings: list[str] = field(default_factory=list)


def _reject_formula_cell(value: Any) -> Any:
    if isinstance(value, str) and value and _FORMULA_PREFIX.match(value.strip()):
        # Do not evaluate — reject formula-looking cells for safety in ingestion path
        raise FormulaRejectedError("formula-like cell content rejected")
    if isinstance(value, str) and len(value) > MAX_CELL_CHARS:
        raise LimitExceededError(f"cell exceeds {MAX_CELL_CHARS} characters")
    return value


def _sanitize_grid(rows: list[list[Any]]) -> list[list[Any]]:
    out: list[list[Any]] = []
    for row in rows:
        if len(row) > MAX_COLUMNS:
            raise LimitExceededError(f"column limit {MAX_COLUMNS} exceeded")
        out.append([_reject_formula_cell(c) for c in row])
    return out


class SpreadsheetReader:
    """Read .xlsx / .csv safely. .xls returns clear UnsupportedFormatError."""

    def read(self, data: bytes, *, filename: str) -> SpreadsheetContent:
        name = (filename or "").lower()
        if name.endswith(".xls") and not name.endswith(".xlsx"):
            raise UnsupportedFormatError(
                "legacy .xls is not supported without a secure xlrd dependency; "
                "convert to .xlsx or .csv"
            )
        if name.endswith(".csv"):
            return self._read_csv(data)
        if name.endswith(".xlsx"):
            return self._read_xlsx(data)
        raise UnsupportedFormatError("unsupported spreadsheet format")

    def _read_xlsx(self, data: bytes) -> SpreadsheetContent:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise UnsupportedFormatError("openpyxl not available") from exc

        # data_only=False — do not evaluate formulas; we reject formula strings
        wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=False)
        try:
            names = list(wb.sheetnames)
            if len(names) > MAX_SHEETS:
                raise LimitExceededError(f"sheet limit {MAX_SHEETS} exceeded")
            sheets: list[SheetPreview] = []
            for sname in names:
                ws = wb[sname]
                rows: list[list[Any]] = []
                scanned = 0
                max_cols = 0
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    scanned += 1
                    if i >= MAX_PREVIEW_ROWS:
                        break
                    cells = list(row)
                    max_cols = max(max_cols, len(cells))
                    rows.append(cells)
                grid = _sanitize_grid(rows) if rows else []
                empty = not any(
                    any(c is not None and str(c).strip() != "" for c in r) for r in grid
                )
                sheets.append(
                    SheetPreview(
                        name=sname,
                        rows=grid,
                        empty=empty,
                        row_count_scanned=scanned,
                        col_count=max_cols,
                    )
                )
            if not sheets:
                raise EmptyFileError("workbook has no sheets")
            return SpreadsheetContent(format="xlsx", sheets=sheets)
        finally:
            wb.close()

    def _read_csv(self, data: bytes) -> SpreadsheetContent:
        encoding, enc_warn = _detect_encoding(data)
        text = data.decode(encoding, errors="replace")
        lines = text.splitlines()
        if len(lines) > MAX_CSV_LINES_SCAN:
            raise LimitExceededError(f"CSV exceeds {MAX_CSV_LINES_SCAN} lines for scan")
        if not lines or all(not ln.strip() for ln in lines):
            raise EmptyFileError("CSV is empty")

        delimiter, delim_warn = _detect_delimiter(lines[:50])
        warnings = [w for w in (enc_warn, delim_warn) if w]

        reader = csv.reader(io.StringIO("\n".join(lines[:MAX_PREVIEW_ROWS])), delimiter=delimiter)
        rows = [list(r) for r in reader]
        grid = _sanitize_grid(rows)
        max_cols = max((len(r) for r in grid), default=0)
        empty = not any(any(str(c).strip() for c in r) for r in grid)
        sheet = SheetPreview(
            name="csv",
            rows=grid,
            empty=empty,
            row_count_scanned=len(lines),
            col_count=max_cols,
        )
        return SpreadsheetContent(
            format="csv",
            sheets=[sheet],
            delimiter=delimiter,
            encoding=encoding,
            warnings=warnings,
        )


def _detect_encoding(data: bytes) -> tuple[str, str | None]:
    """Conservative encoding detection — no silent guess on ambiguity."""
    if data.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig", None
    # Try utf-8 strict
    try:
        data.decode("utf-8")
        return "utf-8", None
    except UnicodeDecodeError:
        pass
    # Try latin-1 only if utf-8 fails (latin-1 always decodes — flag warning)
    try:
        data.decode("cp1252")
        return "cp1252", "encoding_fallback_cp1252"
    except UnicodeDecodeError:
        return "latin-1", "encoding_fallback_latin1_ambiguous"


def _detect_delimiter(sample_lines: list[str]) -> tuple[str, str | None]:
    candidates = [",", ";", "\t"]
    scores: dict[str, list[int]] = {c: [] for c in candidates}
    for line in sample_lines:
        if not line.strip():
            continue
        for c in candidates:
            scores[c].append(line.count(c))
    # Prefer delimiter with stable high count
    best = None
    best_score = -1.0
    for c, counts in scores.items():
        if not counts:
            continue
        avg = sum(counts) / len(counts)
        consistency = 1.0 - (max(counts) - min(counts)) / (max(counts) + 1)
        score = avg * consistency
        if score > best_score:
            best_score = score
            best = c
    if best is None or best_score < 0.5:
        # Ambiguous — default comma but warn (do not silently claim certainty)
        return ",", "delimiter_ambiguous_default_comma"
    # Check near-ties
    ranked = sorted(
        (
            (
                c,
                (sum(counts) / len(counts))
                * (1.0 - (max(counts) - min(counts)) / (max(counts) + 1))
                if counts
                else 0,
            )
            for c, counts in scores.items()
        ),
        key=lambda x: x[1],
        reverse=True,
    )
    warn = None
    if len(ranked) > 1 and ranked[0][1] > 0 and abs(ranked[0][1] - ranked[1][1]) < 0.3:
        warn = "delimiter_near_tie"
    return ranked[0][0], warn
